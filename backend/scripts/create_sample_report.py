"""创建示例 daily report xlsx 供联调下载测试"""
import os
import sys
from pathlib import Path

# 添加项目根路径
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
os.chdir(Path(__file__).resolve().parent.parent)

from openpyxl import Workbook

REPORTS_DIR = Path(os.environ.get("REPORTS_DIR", "reports")) / "daily"
REPORTS_DIR.mkdir(parents=True, exist_ok=True)


def create_sample(dt: str = "2025-02-28"):
    filename = f"daily_report_{dt}.xlsx"
    filepath = REPORTS_DIR / filename
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Report"
    ws["A1"] = "Date"
    ws["B1"] = dt
    ws["A2"] = "Sample"
    ws["B2"] = "Empty report for M0 integration test"
    wb.save(filepath)
    print(f"Created {filepath}")


if __name__ == "__main__":
    create_sample(sys.argv[1] if len(sys.argv) > 1 else "2025-02-28")
